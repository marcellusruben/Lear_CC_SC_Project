# -*- coding: utf-8 -*-
"""
Created on Mon Jun 17 09:19:50 2019

@author: mwinastwan
"""

import numpy as np
from openpyxl import Workbook, load_workbook, styles

def ReadInputFileFromTextFile(filePath):
    
    loadCase = []
    
    with open('InputFile.txt','r') as r:
        
        lines = r.readlines()
        
        for line in lines:
            linelist = line.split()
#            print(linelist)
#            print(len(linelist))
            
            if len(linelist)>1:
                variable = linelist[0]
                value = linelist[1]

                if variable == 'Structure_Variants:':
                    excelName = str(value)
                elif variable == 'Design_Level:':
                    excelName = excelName+'_'+str(value)
                elif variable == 'Investigated_LoadCase:':
                    loadCase = ([x for x in value.split(';')])
                elif variable == 'Welding_Type:':
                    weldingType = str(value)
                elif variable == 'CC_Threshold_EndBeam:':
                    CCEndBeam = float(value)
                elif variable == 'CC_Threshold_MiddleBeam:':
                    CCMiddleBeam = float(value)
                elif variable == 'SC_Threshold_EndBeam:':
                    SCEndBeam = float(value)
                elif variable == 'SC_Threshold_MiddleBeam:':
                    SCMiddleBeam = float(value)
                elif variable == 'Regular_Threshold_EndBeam:':
                    REGEndBeam = float(value)
                elif variable == 'Regular_Threshold_MiddleBeam:':
                    REGMiddleBeam = float(value)
                else:
                    comType = str(value)
                    
    return excelName, loadCase, weldingType, CCEndBeam, CCMiddleBeam, SCEndBeam, SCMiddleBeam, REGEndBeam, REGMiddleBeam, comType

def ReadInputFileFromHyperworks(fileName):
    
    elementID = []
    elementCentroid = []
    componentID = []
    componentName = []
    contourValue = []
    
    with open (fileName+'.txt', 'r') as f:
    
    ### Skip the header ###
        next(f)
    
        Lines = f.readlines()
    
        for line in Lines:
        
            splitLines = line.strip().split(",")

        
            if len(splitLines) > 1:
            
                readLines = splitLines[0:5]
        
                elementID.append(int(readLines[0]))
                elementCentroid.append(readLines[1])
                componentID.append(int(readLines[2]))
                componentName.append(str(readLines[3]))
                contourValue.append(float(readLines[4]))
                
    return elementID, elementCentroid, componentID, componentName, contourValue

def SplitElementCentroid (elementCentroid):
    
    xCoordinate = []
    yCoordinate = []
    zCoordinate = []

    
    for i in range (0, len(elementCentroid)):
    
        xCoordinate.append(float(elementCentroid[i].split()[0]))
        yCoordinate.append(float(elementCentroid[i].split()[1]))
        zCoordinate.append(float(elementCentroid[i].split()[2]))
        
    return xCoordinate, yCoordinate, zCoordinate
    
### ----------------------------------------------------- compute the average value of element centroid ----------------------------------------------

    

def MatchElementCentroid(fileName, xCoordinate, yCoordinate, zCoordinate):
    
    sortedCoordinate = []
    sortedxCoordinate = []
    sortedyCoordinate = []
    sortedzCoordinate = []
    sortedIndex = []
    
    with open(fileName,'r') as r:
        lines = r.readlines()
        
        for line in lines:
            
            sortedCoordinate.append(line)
        
#    print(len(sortedCoordinate))  
    OriginalxCoordinate, OriginalyCoordinate, OriginalzCoordinate = SplitElementCentroid(sortedCoordinate)
    
    count=0
    for i in range (len(OriginalxCoordinate)):
        count=0
        for j in range (len(xCoordinate)):
            
            if xCoordinate[j] == OriginalxCoordinate[i] and yCoordinate[j] == OriginalyCoordinate[i] and zCoordinate[j] == OriginalzCoordinate[i]:
                sortedxCoordinate.append(xCoordinate[j])
                sortedyCoordinate.append(yCoordinate[j])
                sortedzCoordinate.append(zCoordinate[j])
                sortedIndex.append(count)
                break
                
            else:
                count+=1

    return sortedIndex, sortedxCoordinate, sortedyCoordinate, sortedzCoordinate 


    
def WriteSortedCoordinate(filePath, xCoordinate, yCoordinate, zCoordinate):
    
    with open(filePath+'\Coordinate.txt','w') as filewrite:
        
        for i in range (len(xCoordinate)):
            filewrite.write(str(xCoordinate[i])+' ')
            filewrite.write(str(yCoordinate[i])+' ')
            filewrite.write(str(zCoordinate[i])+'\n')
    
def WriteSortedVariableIntoTextFile(elementID, xCoord, yCoord, zCoord, contourValue, componentName):
    
    with open ('sortedVariable.txt','w') as writeFile:
        for i in range (0, len(elementID)):
            writeFile.write(str(elementID[i])+' ')
            writeFile.write(str(xCoord[i])+' ')
            writeFile.write(str(yCoord[i])+' ')
            writeFile.write(str(zCoord[i])+' ')
            writeFile.write(str(contourValue[i])+' ')
            writeFile.write(str(componentName[i])+'\n')

def DetermineListofWeldSeaminOneStructure(elementID, contourValue, valueXDifference, valueYDifference, valueZDifference):
    
    ElList =[]
    ContourList=[]
    numberOfElementOfOneWeldSeam = []
    countElementWeldSeam = 1

    for i in range (len(valueXDifference)):
    
        if valueXDifference[i] <= 4 and valueYDifference[i] <= 4 and valueZDifference[i] <= 4:
        
            countElementWeldSeam += 1
        
        else:
        
            numberOfElementOfOneWeldSeam.append(countElementWeldSeam)
            countElementWeldSeam = 1
        
    ## ADD THE NUMBER OF ELEMENT IN THE END OF THE LOOP
    numberOfElementOfOneWeldSeam.append(countElementWeldSeam) 
#    print(numberOfElementOfOneWeldSeam)
    beamCounter =[]
    for i in range (0,len(numberOfElementOfOneWeldSeam)):
    
        co = 1
    
        for j in range (0,numberOfElementOfOneWeldSeam[i]):
        
        
            if j < numberOfElementOfOneWeldSeam[i]:
            
                beamCounter.append(co) 
                co += 1

    for i in range (0,len(numberOfElementOfOneWeldSeam)):
    
        ElList.append([])
        ContourList.append([])

    weldSeamCounter = 0
    for i in range (0,len(beamCounter)):
    
        counter = 0 + weldSeamCounter
    
        if i == 0:
            ElList[counter].append(elementID[i]) 
            ContourList[counter].append(contourValue[i]) 
        else:
            if beamCounter[i] != 1:
                ElList[counter].append(elementID[i])
                ContourList[counter].append(contourValue[i]) 
            else:
                weldSeamCounter += 1
                ElList[weldSeamCounter].append(elementID[i])
                ContourList[weldSeamCounter].append(contourValue[i])
    
    return ElList, ContourList

def CCandSCDefinition(weldSeamType, CCEnd, CCMiddle, SCEnd, SCMiddle, REGEnd, REGMiddle, ContourList, comType):
    
    valueEndBeam=[]
    setMiddleBeam=[]
    valueMiddleBeam=[]
    resultStatus=[]
    
    print(len(ContourList))
    for i in range (len(ContourList)):
        
       
        valueEndBeam.append(max(ContourList[i][0],ContourList[i][-1]))
        
        setMiddleBeam.append(ContourList[i][1:-1])
            
        valueMiddleBeam.append(max(setMiddleBeam[i]))

#            
    if (weldSeamType == 'Laser' and comType == 'Default') or (weldSeamType == 'Laser' and comType == 'Custom') or (weldSeamType == 'MAG' and comType == 'Custom'):
  
        for i in range (len(valueEndBeam)):
    
            if valueEndBeam[i] > SCEnd or valueMiddleBeam[i] > SCMiddle: ##In this section, insert the threshold value based on user Input
                definition='CC'
            elif valueEndBeam[i] > REGEnd or valueMiddleBeam[i] > REGMiddle: ##In this section, insert the threshold value based on user Input
                definition='SC'
            else:
                definition='REG'
        
            resultStatus.append(definition)
    
    else:
        
        sumEnd = sum(valueEndBeam[i] for i in range (len(valueEndBeam)) if i >0.9)
        sumMiddle = sum(valueMiddleBeam[i] for i in range (len(valueMiddleBeam)) if i >0.9)
        lenEnd = len(sumEnd)
        lenMid = len(sumMiddle)
        
        CCEnd = sumEnd/lenEnd
        CCMiddle = sumMiddle/lenMid
        SCEnd = CCEnd-(CCEnd*0.125)
        SCMiddle = CCMiddle-(CCMiddle*0.125)
        REGEnd = SCEnd-(SCEnd*0.125)
        REGMiddle = SCMiddle-(SCMiddle*0.125)
        
        for i in range (len(valueEndBeam)):
    
            if valueEndBeam[i] > SCEnd or valueMiddleBeam[i] > SCMiddle: ##In this section, insert the threshold value based on user Input
                definition='CC'
            elif valueEndBeam[i] > REGEnd or valueMiddleBeam[i] > REGMiddle: ##In this section, insert the threshold value based on user Input
                definition='SC'
            else:
                definition='REG'
        
            resultStatus.append(definition)
        
        
#    print(resultStatus)    
    return valueEndBeam, valueMiddleBeam, resultStatus


def SaveResultToExcel (fileName, loadCase, column, row, componentName, variable, value, CCEnd, CCMiddle, SCEndBeam, SCMiddleBeam, REGEndBeam, REGMiddleBeam):
    
 
#if elementIDLocation[i] == elementIDLocation[0]: #and ITS ALSO THE FIRST LOOP: 

    wb = Workbook()
    ws = wb.active
    ws.title = str(componentName)

#Cell and row in openpyxl starts from 1, not 0!!!##
#    columnCounter=1
#    rowCounter=2

    thresholdVariable=['','HigherThresholdCC','LowerThresholdCC','HigherThresholdSC','LowerThresholdSC','Regular']
    thresholdValueEnd=['ValueEndBeam','>'+str(SCEndBeam), str(SCEndBeam), str(SCEndBeam), str(REGEndBeam), str(REGEndBeam) ]
    thresholdValueMiddle=['ValueMiddleBeam','>'+str(SCMiddleBeam), str(SCMiddleBeam), str(SCMiddleBeam), str(REGMiddleBeam), str(REGMiddleBeam) ]
   
    for i in range (len(thresholdVariable)):
        ws.cell(row=row,column=column+i).value = thresholdVariable[i]
        ws.cell(row=row,column=column+i).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FFFF00'))
    
    row+=1
    for i in range (len(thresholdVariable)):
        ws.cell(row=row,column=column+i).value = thresholdValueEnd[i]
        ws.cell(row=row,column=column+i).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FFFF00'))

    row+=1
    
    for i in range (len(thresholdVariable)):
        ws.cell(row=row,column=column+i).value = thresholdValueMiddle[i]
        ws.cell(row=row,column=column+i).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FFFF00'))

    row = row+2
    
    ws.cell(row=row,column=column).value = loadCase
    
    column+=1
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18 

    rowIncrement = 0
    for i in range (len(variable)):
    
        ws.cell(row=row+rowIncrement,column=column).value = variable[i]
        rowIncrement+=1
    
    columnIncrement=0
    rowIncrement=0
    count=0
    
    for j in range (len(value)):
        for i in range (len(value[j])):
            columnIncrement+=1
            columnToUse= column+columnIncrement
            count+=1
       
            ws.cell(row=row+rowIncrement,column=columnToUse).value = str(value[j][i])
       
            if value[-1][i] =='CC':
               
              ws.cell(row=row+3,column=columnToUse).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FF0000'))
             
            elif value[-1][i] == 'SC':
              ws.cell(row=row+3,column=columnToUse).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FFFF00'))
               
            else:
              ws.cell(row=row+3,column=columnToUse).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='0000FF00'))
              
        if count == len(value[j]):
           
           lastColumn=columnToUse 
           count=0
           columnToUse=column
           columnIncrement=0
           
        rowIncrement+=1


    wb.save(filename=str(fileName)+'.xlsx')
    
    with open('ColumnLength.txt', 'w')as f:
        f.write('LastColumn'+str(componentName)+':'+' '+str(lastColumn)+'\n')


def RewriteExcelWithoutSavingRow (fileName, loadCase, column, row, componentName, variable, value, CCEnd, CCMiddle, SCEndBeam, SCMiddleBeam, REGEndBeam, REGMiddleBeam):
    
    wb_re_read = load_workbook(filename= str(fileName)+'.xlsx')
    wb = wb_re_read.create_sheet(str(componentName))
    
    thresholdVariable=['','HigherThresholdCC','LowerThresholdCC','HigherThresholdSC','LowerThresholdSC','Regular']
    thresholdValueEnd=['ValueEndBeam','>'+str(SCEndBeam), str(SCEndBeam), str(SCEndBeam), str(REGEndBeam), str(REGEndBeam) ]
    thresholdValueMiddle=['ValueMiddleBeam','>'+str(SCMiddleBeam), str(SCMiddleBeam), str(SCMiddleBeam), str(REGMiddleBeam), str(REGMiddleBeam) ]
   
    for i in range (len(thresholdVariable)):
        wb.cell(row=row,column=column+i).value = thresholdVariable[i]
        wb.cell(row=row,column=column+i).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FFFF00'))
    
    row+=1
    for i in range (len(thresholdVariable)):
        wb.cell(row=row,column=column+i).value = thresholdValueEnd[i]
        wb.cell(row=row,column=column+i).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FFFF00'))

    row+=1
    
    for i in range (len(thresholdVariable)):
        wb.cell(row=row,column=column+i).value = thresholdValueMiddle[i]
        wb.cell(row=row,column=column+i).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FFFF00'))

    row = row+2

    wb.cell(row=row,column=column).value = loadCase
    
    column+=1
    
    wb.column_dimensions['A'].width = 18
    wb.column_dimensions['B'].width = 20
    wb.column_dimensions['C'].width = 18
    wb.column_dimensions['D'].width = 18
    wb.column_dimensions['E'].width = 18    

    rowIncrement = 0
    for i in range (len(variable)):
    
        wb.cell(row=row+rowIncrement,column=column).value = variable[i]
        rowIncrement+=1
    
    columnIncrement=0
    rowIncrement=0
    count=0
    for j in range (len(value)):
        for i in range (len(value[j])):
            columnIncrement+=1
            columnToUse= column+columnIncrement
            count+=1
       
            wb.cell(row=row+rowIncrement,column=columnToUse).value = str(value[j][i])
       
            if value[-1][i] =='CC':
              wb.cell(row=row+3,column=columnToUse).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FF0000'))
             
            elif value[-1][i] == 'SC':
              wb.cell(row=row+3,column=columnToUse).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FFFF00'))
               
            else:
              wb.cell(row=row+3,column=columnToUse).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='0000FF00'))
       
        if count == len(value[j]):
            
           lastColumn=columnToUse  
           count=0
           columnToUse = column
           columnIncrement=0
           
        rowIncrement+=1

    wb_re_read.save(filename=str(fileName)+'.xlsx')
    
    with open('ColumnLength.txt', 'a')as f:
        f.write('LastColumn'+str(componentName)+':'+' '+str(lastColumn)+'\n')
    
        
def RewriteExcel (fileName, loadCase, column, row, componentName, variable, value, CCEnd, CCMiddle, SCEndBeam, SCMiddleBeam, REGEndBeam, REGMiddleBeam):
    
    wb_re_read = load_workbook(filename= str(fileName)+'.xlsx')
    wb = wb_re_read.create_sheet(str(componentName))
    
    thresholdVariable=['','HigherThresholdCC','LowerThresholdCC','HigherThresholdSC','LowerThresholdSC','Regular']
    thresholdValueEnd=['ValueEndBeam','>'+str(SCEndBeam), str(SCEndBeam), str(SCEndBeam), str(REGEndBeam), str(REGEndBeam) ]
    thresholdValueMiddle=['ValueMiddleBeam','>'+str(SCMiddleBeam), str(SCMiddleBeam), str(SCMiddleBeam), str(REGMiddleBeam), str(REGMiddleBeam) ]
   
    for i in range (len(thresholdVariable)):
        wb.cell(row=row,column=column+i).value = thresholdVariable[i]
        wb.cell(row=row,column=column+i).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FFFF00'))
    
    row+=1
    for i in range (len(thresholdVariable)):
        wb.cell(row=row,column=column+i).value = thresholdValueEnd[i]
        wb.cell(row=row,column=column+i).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FFFF00'))

    row+=1
    
    for i in range (len(thresholdVariable)):
        wb.cell(row=row,column=column+i).value = thresholdValueMiddle[i]
        wb.cell(row=row,column=column+i).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FFFF00'))

    row = row+2

    wb.cell(row=row,column=column).value = loadCase
    
    column+=1
    
    wb.column_dimensions['A'].width = 18
    wb.column_dimensions['B'].width = 20
    wb.column_dimensions['C'].width = 18
    wb.column_dimensions['D'].width = 18
    wb.column_dimensions['E'].width = 18 

    rowIncrement = 0
    for i in range (len(variable)):
    
        wb.cell(row=row+rowIncrement,column=column).value = variable[i]
        rowIncrement+=1
    
    columnIncrement=0
    rowIncrement=0
    count=0
    for j in range (len(value)):
        for i in range (len(value[j])):
            columnIncrement+=1
            columnToUse= column+columnIncrement
            count+=1
       
            wb.cell(row=row+rowIncrement,column=columnToUse).value = str(value[j][i])
       
            if value[-1][i] =='CC':
              wb.cell(row=row+3,column=columnToUse).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FF0000'))
             
            elif value[-1][i] == 'SC':
              wb.cell(row=row+3,column=columnToUse).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FFFF00'))
               
            else:
              wb.cell(row=row+3,column=columnToUse).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='0000FF00'))
       
        if count == len(value[j]):
            
           lastColumn=columnToUse 
           count=0
           columnToUse = column
           lastRow=row+rowIncrement
           columnIncrement=0
           
        rowIncrement+=1

    wb_re_read.save(filename=str(fileName)+'.xlsx')
    
    with open('ColumnLength.txt', 'a')as f:
        f.write('LastColumn'+str(componentName)+':'+' '+str(lastColumn)+'\n')
    
    with open('RowandColumnInfo.txt', 'w')as f:
        f.write('LastRow:'+' '+str(lastRow)+'\n')
        f.write('FirstColumn:'+' '+str(column)+'\n')


    with open('RowContainingResult.txt', 'w')as f:
        f.write(str(loadCase)+'_'+str(componentName)+':'+' '+str(lastRow)+'\n')
        
def LoopWithoutSavingLatestRow (fileName, loadCase, column, row, componentName, variable, value):
    
    wb_re_read = load_workbook(filename= str(fileName)+'.xlsx')
    wb = wb_re_read.get_sheet_by_name(str(componentName))

    wb.cell(row=row,column=column).value = loadCase
    
    column+=1
    wb.column_dimensions['B'].width = 20
    rowIncrement = 0
    for i in range (len(variable)):
    
        wb.cell(row=row+rowIncrement,column=column).value = variable[i]
        rowIncrement+=1
    
    columnIncrement=0
    rowIncrement=0
    count=0
    for j in range (len(value)):
        for i in range (len(value[j])):
            columnIncrement+=1
            columnToUse= column+columnIncrement
            count+=1
       
            wb.cell(row=row+rowIncrement,column=columnToUse).value = str(value[j][i])
       
            if value[-1][i] =='CC':
              wb.cell(row=row+3,column=columnToUse).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FF0000'))
             
            elif value[-1][i] == 'SC':
              wb.cell(row=row+3,column=columnToUse).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FFFF00'))
               
            else:
              wb.cell(row=row+3,column=columnToUse).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='0000FF00'))
       
        if count == len(value[j]):
            
           count=0
           columnToUse = column
           
           columnIncrement=0
           
        rowIncrement+=1

    wb_re_read.save(filename=str(fileName)+'.xlsx')


def LoopWithSavingLatestRow (fileName, loadCase, column, row, componentName, variable, value):
    
    wb_re_read = load_workbook(filename= str(fileName)+'.xlsx')
    wb = wb_re_read.get_sheet_by_name(str(componentName))

    wb.cell(row=row,column=column).value = loadCase
    
    column+=1
    wb.column_dimensions['B'].width = 20
    rowIncrement = 0
    for i in range (len(variable)):
    
        wb.cell(row=row+rowIncrement,column=column).value = variable[i]
        rowIncrement+=1
    
    columnIncrement=0
    rowIncrement=0
    count=0
    for j in range (len(value)):
        for i in range (len(value[j])):
            columnIncrement+=1
            columnToUse= column+columnIncrement
            count+=1
       
            wb.cell(row=row+rowIncrement,column=columnToUse).value = str(value[j][i])
       
            if value[-1][i] =='CC':
              wb.cell(row=row+3,column=columnToUse).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FF0000'))
             
            elif value[-1][i] == 'SC':
              wb.cell(row=row+3,column=columnToUse).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FFFF00'))
               
            else:
              wb.cell(row=row+3,column=columnToUse).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='0000FF00'))
       
        if count == len(value[j]):
           
           
           count=0
           columnToUse = column
           lastRow= row+rowIncrement
           columnIncrement=0
           
        rowIncrement+=1

    wb_re_read.save(filename=str(fileName)+'.xlsx')
    
    with open('RowandColumnInfo.txt', 'w')as f:
        f.write('LastRow:'+' '+str(lastRow)+'\n')
        f.write('FirstColumn:'+' '+str(column)+'\n')


    with open('RowContainingResult.txt', 'a')as f:
        f.write(str(loadCase)+'_'+str(componentName)+':'+' '+str(lastRow)+'\n')
        
def ReadLastRow(filePath):
    
    with open(filePath+'/RowandColumnInfo.txt','r') as r:
        lines = r.readlines()
        
        for line in lines:
            linelist = line.split()
            
            if linelist[0]=='LastRow:':
                lastRow=linelist[1]
            elif linelist[0] == 'FirstColumn:':
                firstColumn=linelist[1]
            
    
    return lastRow, firstColumn

def ComputeFinalStatusExcel(fileName, componentName, columnLength, row, column, resultRow):
    
    wb_re_read = load_workbook(filename= str(fileName)+'.xlsx')
    wb = wb_re_read.get_sheet_by_name(str(componentName))
    resultList = []
    wb.cell(row=row,column=column-1).value = 'Final Status'
    
    for i in range (columnLength-column+1):
        
        for j in range (len(resultRow)):
            
            if wb.cell(row=int(resultRow[j]),column=column+i).value == 'CC':
                wb.cell(row=row, column=column+i).value= 'CC'
                
            else:
                if wb.cell(row=int(resultRow[j]),column=column+i).value == 'SC':
                    wb.cell(row=row, column=column+i).value= 'SC'
                else:
                    wb.cell(row=row, column=column+i).value= 'REG'
                    
        resultList.append(wb.cell(row=row, column=column+i).value)
            
        if wb.cell(row=row, column=column+i).value == 'CC':
            wb.cell(row=row, column=column+i).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FF0000'))
        elif wb.cell(row=row, column=column+i).value == 'SC':
            wb.cell(row=row, column=column+i).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FFFF00'))
        else:
            wb.cell(row=row, column=column+i).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='0000FF00'))
    
    amountCC = resultList.count('CC')
    amountSC = resultList.count('SC')
    amountREG = resultList.count('REG')
    countVariable= ['CC','SC','Regular']
    countValue= [str(amountCC), str(amountSC), str(amountREG)]
                 
    row+=2
    wb.cell(row=row,column=column-1).value = 'Status Count'
    
    row+=1
    for i in range (len(countVariable)):
        wb.cell(row=row+i, column=column-1).value= countVariable[i]
        wb.cell(row=row+i, column=column-1).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FFFF00'))
        wb.cell(row=row+i, column=column).value= countValue[i]
        wb.cell(row=row+i, column=column).fill= styles.fills.PatternFill(patternType='solid', fgColor= styles.colors.Color(rgb='00FFFF00'))    
    
    wb_re_read.save(filename=str(fileName)+'.xlsx')
    